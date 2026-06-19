#!/usr/bin/env python3
"""Import the V3 medication catalog without overwriting existing IAM data."""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import uuid
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_COLUMNS = [
    "Nom_Medicament",
    "Dosage",
    "Concentration",
    "Forme_Galenique",
    "Presentation",
    "Volume",
    "Quantite_Unites",
    "Quantite_Boites",
    "Substances",
    "Specialite",
]


def load_env(path: Path = Path(".env")) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def db_config() -> dict[str, Any]:
    if os.environ.get("MYSQL_URL"):
        from urllib.parse import urlparse

        parsed = urlparse(os.environ["MYSQL_URL"])
        return {
            "host": parsed.hostname or "127.0.0.1",
            "port": parsed.port or 3306,
            "user": parsed.username or "root",
            "password": parsed.password or "",
            "database": parsed.path.lstrip("/") or "projet_ipa",
        }

    return {
        "host": os.environ.get("MYSQLHOST") or os.environ.get("DB_HOST", "127.0.0.1"),
        "port": int(os.environ.get("MYSQLPORT") or os.environ.get("DB_PORT", "3306")),
        "user": os.environ.get("MYSQLUSER") or os.environ.get("DB_USER", "root"),
        "password": os.environ.get("MYSQLPASSWORD") or os.environ.get("DB_PASSWORD", ""),
        "database": os.environ.get("MYSQLDATABASE") or os.environ.get("DB_NAME", "projet_ipa"),
    }


def text_value(value: Any) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    if not value or value.lower() == "nan":
        return None
    return value


def decimal_value(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize(value: str | None) -> str:
    if not value:
        return ""
    value = value.upper()
    value = re.sub(r"[\u0300-\u036f]", "", value)
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def split_substances(value: str | None) -> list[str]:
    if not value:
        return []
    substances = []
    for part in value.split("|"):
        cleaned = text_value(part)
        if cleaned:
            substances.append(cleaned)
    return substances


def row_hash(row: dict[str, Any]) -> str:
    payload = "|".join(str(row.get(col) or "") for col in EXPECTED_COLUMNS)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def normalized_key(row: dict[str, Any]) -> str:
    payload = "|".join(
        normalize(str(row.get(col) or ""))
        for col in ("Nom_Medicament", "Dosage", "Forme_Galenique", "Substances", "Specialite")
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def searchable_text(row: dict[str, Any]) -> str:
    return " ".join(str(row.get(col) or "") for col in ("Nom_Medicament", "Dosage", "Forme_Galenique", "Substances", "Specialite")).strip()


def read_rows(path: Path, limit: int | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header = [text_value(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    missing = [col for col in EXPECTED_COLUMNS if col not in header]
    if missing:
        raise ValueError(f"Missing expected columns: {', '.join(missing)}")

    indexes = {name: header.index(name) for name in EXPECTED_COLUMNS}
    rows = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        record = {
            "row_number": row_number,
            "Nom_Medicament": text_value(cells[indexes["Nom_Medicament"]].value),
            "Dosage": text_value(cells[indexes["Dosage"]].value),
            "Concentration": text_value(cells[indexes["Concentration"]].value),
            "Forme_Galenique": text_value(cells[indexes["Forme_Galenique"]].value),
            "Presentation": text_value(cells[indexes["Presentation"]].value),
            "Volume": text_value(cells[indexes["Volume"]].value),
            "Quantite_Unites": decimal_value(cells[indexes["Quantite_Unites"]].value),
            "Quantite_Boites": decimal_value(cells[indexes["Quantite_Boites"]].value),
            "Substances": text_value(cells[indexes["Substances"]].value),
            "Specialite": text_value(cells[indexes["Specialite"]].value),
        }
        if record["Nom_Medicament"]:
            record["source_hash"] = row_hash(record)
            record["normalized_key"] = normalized_key(record)
            rows.append(record)
        if limit and len(rows) >= limit:
            break
    return rows


def preload_existing(cursor) -> tuple[dict[str, int], dict[str, int]]:
    cursor.execute("SELECT id, specialites FROM specialites")
    specialites = {normalize(name): item_id for item_id, name in cursor.fetchall() if name}
    cursor.execute("SELECT id, substances FROM substances")
    substances = {normalize(name): item_id for item_id, name in cursor.fetchall() if name}
    return specialites, substances


def audit(cursor, batch_id: str, action_type: str, status: str, reason: str, payload: dict[str, Any], catalog_id: int | None = None, target_table: str | None = None, target_id: int | None = None) -> None:
    cursor.execute(
        """
        INSERT INTO medication_enrichment_audit
            (import_batch_id, catalog_id, action_type, target_table, target_id, status, reason, payload_json)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (batch_id, catalog_id, action_type, target_table, target_id, status, reason, json.dumps(payload, ensure_ascii=False)),
    )


def insert_staging(cursor, batch_id: str, source_file: str, row: dict[str, Any]) -> None:
    cursor.execute(
        """
        INSERT INTO medication_catalog_staging
            (import_batch_id, source_file, `row_number`, nom_medicament, dosage, concentration,
             forme_galenique, presentation, volume, quantite_unites, quantite_boites,
             substances, specialite, source_hash, normalized_key)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            nom_medicament = VALUES(nom_medicament),
            dosage = VALUES(dosage),
            concentration = VALUES(concentration),
            forme_galenique = VALUES(forme_galenique),
            presentation = VALUES(presentation),
            volume = VALUES(volume),
            quantite_unites = VALUES(quantite_unites),
            quantite_boites = VALUES(quantite_boites),
            substances = VALUES(substances),
            specialite = VALUES(specialite),
            source_hash = VALUES(source_hash),
            normalized_key = VALUES(normalized_key)
        """,
        (
            batch_id,
            source_file,
            row["row_number"],
            row["Nom_Medicament"],
            row["Dosage"],
            row["Concentration"],
            row["Forme_Galenique"],
            row["Presentation"],
            row["Volume"],
            row["Quantite_Unites"],
            row["Quantite_Boites"],
            row["Substances"],
            row["Specialite"],
            row["source_hash"],
            row["normalized_key"],
        ),
    )


def upsert_catalog(cursor, batch_id: str, row: dict[str, Any], existing_specialites: dict[str, int]) -> int:
    iam_specialite_id = existing_specialites.get(normalize(row["Nom_Medicament"])) or existing_specialites.get(normalize(row["Specialite"]))
    match_status = "specialite_exact" if iam_specialite_id else "unmatched"
    cursor.execute(
        """
        INSERT INTO medication_catalog
            (nom_medicament, dosage, concentration, forme_galenique, presentation, volume,
             quantite_unites, quantite_boites, substances, specialite, searchable_text,
             source_hash, normalized_key, iam_specialite_id, match_status, last_import_batch_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            nom_medicament = VALUES(nom_medicament),
            dosage = VALUES(dosage),
            concentration = VALUES(concentration),
            forme_galenique = VALUES(forme_galenique),
            presentation = VALUES(presentation),
            volume = VALUES(volume),
            quantite_unites = VALUES(quantite_unites),
            quantite_boites = VALUES(quantite_boites),
            substances = VALUES(substances),
            specialite = VALUES(specialite),
            searchable_text = VALUES(searchable_text),
            normalized_key = VALUES(normalized_key),
            iam_specialite_id = COALESCE(medication_catalog.iam_specialite_id, VALUES(iam_specialite_id)),
            match_status = IF(medication_catalog.iam_specialite_id IS NULL, VALUES(match_status), medication_catalog.match_status),
            last_import_batch_id = VALUES(last_import_batch_id)
        """,
        (
            row["Nom_Medicament"],
            row["Dosage"],
            row["Concentration"],
            row["Forme_Galenique"],
            row["Presentation"],
            row["Volume"],
            row["Quantite_Unites"],
            row["Quantite_Boites"],
            row["Substances"],
            row["Specialite"],
            searchable_text(row),
            row["source_hash"],
            row["normalized_key"],
            iam_specialite_id,
            match_status,
            batch_id,
        ),
    )
    cursor.execute("SELECT id FROM medication_catalog WHERE source_hash = %s", (row["source_hash"],))
    return cursor.fetchone()[0]


def enrich_iam(cursor, batch_id: str, row: dict[str, Any], catalog_id: int, specialites: dict[str, int], substances: dict[str, int]) -> dict[str, int]:
    stats = Counter()
    specialite_name = row["Nom_Medicament"]
    specialite_key = normalize(specialite_name)
    if not specialite_name or len(specialite_name) > 45:
        audit(cursor, batch_id, "specialite_insert", "skipped", "Nom_Medicament absent or longer than specialites.specialites varchar(45)", row, catalog_id, "specialites")
        stats["specialite_skipped"] += 1
        specialite_id = None
    elif specialite_key in specialites:
        specialite_id = specialites[specialite_key]
        stats["specialite_existing"] += 1
    else:
        cursor.execute("INSERT INTO specialites (specialites) VALUES (%s)", (specialite_name.upper(),))
        specialite_id = cursor.lastrowid
        specialites[specialite_key] = specialite_id
        audit(cursor, batch_id, "specialite_insert", "applied", "Specialite added from catalog Nom_Medicament", row, catalog_id, "specialites", specialite_id)
        stats["specialite_inserted"] += 1

    for substance in split_substances(row["Substances"]):
        substance_key = normalize(substance)
        if not substance_key:
            continue
        if substance_key in substances:
            substance_id = substances[substance_key]
            stats["substance_existing"] += 1
        else:
            cursor.execute("INSERT INTO substances (substances) VALUES (%s)", (substance.upper(),))
            substance_id = cursor.lastrowid
            substances[substance_key] = substance_id
            audit(cursor, batch_id, "substance_insert", "applied", "Substance added from catalog", {"substance": substance, **row}, catalog_id, "substances", substance_id)
            stats["substance_inserted"] += 1

        if specialite_id:
            cursor.execute(
                "SELECT COUNT(*) FROM liaisons_ss WHERE id_specialites = %s AND id_substance = %s",
                (specialite_id, substance_id),
            )
            exists = cursor.fetchone()[0] > 0
            if exists:
                stats["liaison_existing"] += 1
            else:
                cursor.execute(
                    "INSERT IGNORE INTO liaisons_ss (id_specialites, id_substance) VALUES (%s, %s)",
                    (specialite_id, substance_id),
                )
                if cursor.rowcount:
                    audit(cursor, batch_id, "liaison_ss_insert", "applied", "Specialite-substance link added from catalog", {"substance": substance, **row}, catalog_id, "liaisons_ss", specialite_id)
                    stats["liaison_inserted"] += 1
                else:
                    audit(cursor, batch_id, "liaison_ss_insert", "skipped", "Existing liaisons_ss uniqueness constraint prevented insert", {"substance": substance, **row}, catalog_id, "liaisons_ss", specialite_id)
                    stats["liaison_skipped_constraint"] += 1

    return stats


def report(rows: list[dict[str, Any]]) -> None:
    source_counts = Counter(row["source_hash"] for row in rows)
    key_counts = Counter(row["normalized_key"] for row in rows)
    name_to_keys: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        name_to_keys[normalize(row["Nom_Medicament"])].add(row["normalized_key"])

    exact_duplicates = sum(count - 1 for count in source_counts.values() if count > 1)
    business_duplicates = sum(count - 1 for count in key_counts.values() if count > 1)
    ambiguous_names = sum(1 for keys in name_to_keys.values() if len(keys) > 1)
    substances = {normalize(item) for row in rows for item in split_substances(row["Substances"]) if normalize(item)}

    print("Catalog import report")
    print(f"Rows read: {len(rows)}")
    print(f"Unique source hashes: {len(source_counts)}")
    print(f"Exact duplicate rows: {exact_duplicates}")
    print(f"Business duplicate rows: {business_duplicates}")
    print(f"Ambiguous medication names: {ambiguous_names}")
    print(f"Unique substances in file slice: {len(substances)}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True, help="Path to Lien_SPECIALITES-SUBSTANCES_CLEANED_11-06-2026.xlsx")
    parser.add_argument("--limit", type=int, default=None, help="Maximum number of rows to import")
    parser.add_argument("--dry-run", action="store_true", help="Read and report without writing")
    parser.add_argument("--apply-catalog", action="store_true", help="Write staging and medication_catalog rows")
    parser.add_argument("--apply-iam-enrichment", action="store_true", help="Safely add missing specialites/substances/liaisons_ss")
    parser.add_argument("--report-duplicates", action="store_true", help="Print duplicate summary")
    args = parser.parse_args()

    path = Path(args.file).expanduser()
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        return 2

    load_env()
    rows = read_rows(path, args.limit)
    if args.dry_run or args.report_duplicates or not (args.apply_catalog or args.apply_iam_enrichment):
        report(rows)
    if args.dry_run or not (args.apply_catalog or args.apply_iam_enrichment):
        return 0

    batch_id = str(uuid.uuid4())
    import mysql.connector

    conn = mysql.connector.connect(**db_config())
    cursor = conn.cursor()
    specialites, substances = preload_existing(cursor)
    stats = Counter()

    try:
        for row in rows:
            if args.apply_catalog:
                insert_staging(cursor, batch_id, path.name, row)
                catalog_id = upsert_catalog(cursor, batch_id, row, specialites)
                stats["catalog_upserted"] += 1
            else:
                cursor.execute("SELECT id FROM medication_catalog WHERE source_hash = %s", (row["source_hash"],))
                found = cursor.fetchone()
                catalog_id = found[0] if found else None

            if args.apply_iam_enrichment:
                if catalog_id is None:
                    audit(cursor, batch_id, "iam_enrichment", "skipped", "Catalog row does not exist; run --apply-catalog first", row)
                    stats["iam_skipped_missing_catalog"] += 1
                else:
                    stats.update(enrich_iam(cursor, batch_id, row, catalog_id, specialites, substances))

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()

    print(f"Import batch: {batch_id}")
    for key in sorted(stats):
        print(f"{key}: {stats[key]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
